from io import BytesIO

from django.http import QueryDict
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from openpyxl import load_workbook

from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.authentication import SessionAuthentication, BasicAuthentication

from dataView.models import Country
from dataView.models import Department
from dataView.models import Product
from dataView.models import Sale



class ByPassSessionAuthentication(SessionAuthentication):
    """
    We use this to by-pass token authentification.
    Not Ideal.
    """

    def enforce_csrf(self, request):
        return


class BaseView(APIView):
    """
    Parent class for all views.
    defines the methods needed to create an API.
    """

    def request_data_to_dict(self, *args, **kwargs):
        if isinstance(self.request.data, QueryDict):
            data = self.request.data.dict()

        else:
            data = self.request.data

        for key, value in kwargs.items():
            if value == "null":
                kwargs[key] = None
        data.update(kwargs)
        self.data = data

    def process_request(self):
        pass

    def get_response(self, request, *args, **kwargs):
        self.request = request
        self.request_data_to_dict(*args, **kwargs)
        response_data = self.process_request()
        return self.create_response(response_data)

    def create_response(self, response_data):
        if not response_data:
            response_data = dict()
        return JsonResponse(dict(response_data=response_data))
    

@method_decorator(csrf_exempt, name='dispatch')
class FileUploadView(BaseView):
    """
    Receives a file upload from the client side.
    Might work well for now but given that it receives the whole file,
    it could be improved by reading a file directly from google drive/ s3 etc
    """
    parser_classes = [MultiPartParser, FormParser]
    authentication_classes = (ByPassSessionAuthentication, BasicAuthentication)

    def post(self, request):
        return self.get_response(request)

    def process_request(self):
        for filename, file in self.data.items():
            book = load_workbook(filename=BytesIO(file.file.read()), read_only=True)
            sheet = book.worksheets[0]
            for record in self.__read_xlsx_data(sheet):
                self.__save_sale(record)
            book.close()
        return {'status': 'File upload received'}

    def __read_xlsx_data(self, sheet):
        headers = None
        for row in sheet.iter_rows():
            cells = [self.__remove_spaces_from_strings(x.value) for x in row]
            if not any(cells):
                break
            if not headers:
                headers = cells
            else:
                yield dict(zip(headers, cells))

    def __remove_spaces_from_strings(self, value):
        if isinstance(value, str):
            value = value.lower().strip()
        return value

    def __save_sale(self, record):
        department = record['department']
        country = record['country']
        product = record['product']
        manufacturing_cost = record['manufacturing price']
        discount_band = record['discount band']
        discount = record['discounts']
        quantity = record['units sold']
        price = record['sale price']
        sale = record['sales']
        profit = record["profit"]
        cogs = record['cogs']
        date = record['date']

        sale = Sale.add_sale(
            department, country, product, manufacturing_cost,
            discount_band, discount, quantity, price, sale, profit, cogs, date
            )


class SaleView(BaseView):
    def get(self, request, *args, **kwargs):
        return self.get_response(request, *args, **kwargs)

    def process_request(self):
        print(self.data)
        sales = Sale.get_all(
            department=self.data.get('department'),
            country=self.data.get('country'),
            product=self.data.get('product'),
            before=self.data.get('from_date'),
            after=self.data.get('to_date')
            )

        sales_json = [sale.json() for sale in sales]
        return sales_json


class ChartView(BaseView):

    def get(self, request, *args, **kwargs):
        return self.get_response(request, *args, **kwargs)

    def process_request(self):
        from dataView.dummydata import dummydata
        chart_data = []

        data = Sale.get_chart_data(
            department=self.data.get('department'),
            country=self.data.get('country'),
            product=self.data.get('product'),
            before=self.data.get('from_date'),
            after=self.data.get('to_date'))

        for entry in data:
            entry['cogs'] = float(entry['cogs'])
            entry['sales'] = float(entry['sales'])
            entry['profit'] = float(entry['profit'])
            chart_data.append(entry)

        return chart_data


class DepartmentView(BaseView):

    def get(self, request):
        return self.get_response(request)

    def process_request(self):
        return Department.get_all_as_json()


class CountryView(BaseView):

    def get(self, request):
        return self.get_response(request)

    def process_request(self):
        return Country.get_all_as_json()


class ProductView(BaseView):

    def get(self, request):
        return self.get_response(request)

    def process_request(self):
        return Product.get_all_as_json()
        